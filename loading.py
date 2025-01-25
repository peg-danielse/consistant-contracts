import openpyxl

def read_xlsx_to_lists(file_path):
    questions = []
    answers = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        print()

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:  # Ensure there are at least two columns with data
                questions.append(row[0])
                answers.append(row[1])

        return questions, answers

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")


a, b = read_xlsx_to_lists("./prompts_corrected.xlsx")

print(a)