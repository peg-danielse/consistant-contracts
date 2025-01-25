import sys
import os
import openpyxl
import pandas as pd

from typing import Optional
import fire
from llama_models.llama3.api.datatypes import RawMessage, StopReason
from llama_models.llama3.reference_impl.generation import Llama


def generate_context():

    #dialog = [RawMessage(role="user", content="I am going to give you a document. Read it and understand it. Answer all following questions with the document in mind."),
    #    RawMessage(role="user", content="This is a document. It is a legal document but is completely empty and does not make any legal statements.")]
    
    #return dialog
    print(os.getcwd())
    context_documents = ["dora-chapter-V-art-28.txt"]#, "dora-chapter-V-art-30.txt", "dora-chapter-V-art-30.txt"] 
    #test_documents = ["bench-Atlas-CA.txt"]#,"bench-Micro-SA.txt","bench-SG-MSA.txt","bench-TSP.txt","bench-TTSP.txt"]
    test_documents = ["bench-Atlas-CA.txt"]
    for i, fn in enumerate(context_documents):
        fn = "llama_models/scripts/" + fn
        with open(fn, "r") as file:
            context_documents[i] = file.read()

    for i, fn in enumerate(test_documents):
        fn = "llama_models/scripts/" + fn
        with open(fn, "r") as file:
            test_documents[i] = file.read()

    system_add_context_message = "add the following document to your context."
    system_add_test_message = "add the following document to your context and only answer questions about this document."

    dialog = [
        RawMessage(role="system", content=system_add_context_message),
        RawMessage(role="system", content=context_documents[0]),
        RawMessage(role="system", content=system_add_test_message),
        RawMessage(role="system", content=test_documents[0])]
    
    return dialog

def generateParaphrasePrompt(method, sentence):
    prompt = f""":
    Today I want you to learn the ways of paraphrasing a sentence. Below are few methods with examples. Go through
    them carefully.
    1. Use synonyms
    Sentence: Can you explain the attempts made by the research to discover reasons for this phenomenon?
    Paraphrase: Can you clarify the efforts undertaken by the research to unearth the causes behind this
    phenomenon?
    2. Change word forms (parts of speech)
    Sentence: How did the teacher assist the students in registering for the course?
    Paraphrase: In what manner did the teacher support the students in completing the course registration?
    3. Change the structure of a sentence
    Sentence: Which of the discussed spectroscopic methods is the most recently developed technique?
    Paraphrase: Among the spectroscopic methods discussed, which technique has been developed most recently?
    4. Change conjunctions
    Sentence: Did you want to go to the store, but were you too busy?
    Paraphrase: Although you were busy, did you still want to go to the store?
    Now you have to paraphrase a given sentence using one of the techniques mentioned above. I will provide you
    the number of the technique to use.
    Technique Number: {method}
    Sentence: {sentence}
    Paraphrase (Do not give any preamble, just give the paraphrased text):"""

    return prompt

def generateRankPrompt(question, answers):

    prompt = f"""
        Question: {question}
        For the question above there are several options given, choose one among them which seems to be the most
        correct. Do not give any other output. Only the number.
        Option 1: {answers[0]}
        Option 2: {answers[1]}
        Option 3: {answers[2]}
        Option 4: {answers[3]}
        Option 5: {answers[4]}
        Option 6: Don’t know the correct answer
        Answer:
        """
    
    return prompt

def read_xlsx_to_lists(file_path):
    questions = []
    answers = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:  # Ensure there are at least two columns with data
                questions.append(row[0])
                answers.append(row[1])

        return questions, answers

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_reply(generator, context, prompt, max_gen_len, temperature, top_p):
    #print("CONTEXT -> ", context)
    if context:
        context.append(RawMessage(role="user", content=prompt))
        model_input = context
    else:
        model_input = [RawMessage(role="user", content=prompt)]

    result = generator.chat_completion(model_input, max_gen_len=max_gen_len, temperature=temperature, top_p=top_p)
    
    #print("\n\n=== MODEL PROMPT ===")
    #for msg in model_input:
    #    print(f"{msg.role.capitalize()}: {msg.content}\n")

    out_message = result.generation
    #print("\n\n=== MODEL OUTPUT ===")
    #print(f"> {out_message.role.capitalize()}: {out_message.content}")
    #print("\n==================================\n")

    return out_message.content
    

def run_main(ckpt_dir: str, temperature: float = 0.6, top_p: float = 0.9, max_seq_len: int = 15000, max_batch_size: int = 4, max_gen_len: Optional[int] = None, model_parallel_size: Optional[int] = None):

    generator = Llama.build(
        ckpt_dir=ckpt_dir,
        max_seq_len=max_seq_len,
        max_batch_size=max_batch_size,
        model_parallel_size=model_parallel_size,
        seed=42
    )
    
    questions, answers = read_xlsx_to_lists("llama_models/scripts/article_28_question_answer_pairs.xlsx")
    context_input = generate_context()

    final_answers = []
    
    #print("\n\n=== ALL QUESTIONS ===")
    #print(questions)
    #print(len(questions))
    for question_i in range(6,len(questions)):
        print("\n\n=== ITERATION", question_i,"===")
        print("\n\n === INPUT PROMPT ===")
        #questions[0] = questions[0] #+ "\n\n(LIMIT YOUR ANSWER TO 500 WORDS MAXIMUM)"
        print(questions[question_i])
        #print("========")
        #print(answers[0])

        print("\n\n=== ORIGINAL ANSWER ===")
        original_answer = generate_reply(generator, context_input, questions[0], max_gen_len, temperature, top_p)
        #print(original_answer)

        paraphrasedQuestions = []
        for i in range(1,5):
            paraphrasedQuestion = generate_reply(generator, None, generateParaphrasePrompt(i,questions[0]), max_gen_len, temperature, top_p)
            paraphrasedQuestions.append(paraphrasedQuestion)

        print("\n\n=== PARAPHRASED QUESTIONS ===")
        for paraphrased_question in paraphrasedQuestions:
            print("\n === Question ===")
            #print("\n", paraphrased_question)

        paraphrased_answers = []
        for question in paraphrasedQuestions:
            paraphrased_answer = generate_reply(generator, context_input, question, max_gen_len, temperature, top_p)
            paraphrased_answers.append(paraphrased_answer)

        print("\n\n=== PARAPHRASED ANSWERS ===")
        for paraphrased_answer in paraphrased_answers:
            print("\n === ANSWER ===")
            #print(paraphrased_answer)

        all_answers = [original_answer] + paraphrased_answers
        rankPrompt = generateRankPrompt(questions, all_answers)
        #print("RANK PROMPT ->", rankPrompt)
        
        print("\n\n=== BEST RANKED ANSWER ===")
        best_answer = generate_reply(generator, None, rankPrompt, max_gen_len, temperature, top_p)
        print(best_answer)
        print(int(best_answer))

        print(all_answers[int(best_answer)])

        final_answers.append(all_answers[int(best_answer)])
    print("=== FINAL ANSWERS ===")    
    print(final_answers)
    #print(generate_reply(generator, generate_context(),"Given the document info, what is the name of the cat?", max_gen_len, temperature, top_p))

    data = {"question": questions, "answer": final_answers}
    df = pd.DataFrame(data)

    # Save to a CSV file
    df.to_csv("atlas_output.csv", index=False)


# Example usage
if __name__ == "__main__":
    print("hello test")
    
    fire.Fire(run_main)

    quit()
    

