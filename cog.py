import sys
import openpyxl

from typing import Optional
import fire
from llama_models.llama3.api.datatypes import RawMessage, StopReason
from llama_models.llama3.reference_impl.generation import Llama


def generateParaphrasePrompt(method, sentence):
    prompt = f""":
    Today I want you to learn the ways of paraphrasing a sentence. Below are few methods with examples. Go through
    them carefully.
    1. Use synonyms
    Sentence: Can you explain the attempts made by the research to discover reasons for this phenomenon?
    Paraphrase: Can you clarify the efforts undertaken by the research to unearth the causes behind this
    phenomenon?
    2. Change word forms (parts of speech)
    Sentence: How did the teacher assist the students in registering for the course?
    Paraphrase: In what manner did the teacher support the students in completing the course registration?
    3. Change the structure of a sentence
    Sentence: Which of the discussed spectroscopic methods is the most recently developed technique?
    Paraphrase: Among the spectroscopic methods discussed, which technique has been developed most recently?
    4. Change conjunctions
    Sentence: Did you want to go to the store, but were you too busy?
    Paraphrase: Although you were busy, did you still want to go to the store?
    Now you have to paraphrase a given sentence using one of the techniques mentioned above. I will provide you
    the number of the technique to use.
    Technique Number: {method}
    Sentence: {sentence}
    Paraphrase (Do not give any preamble, just give the paraphrased text):"""

    return prompt

def read_xlsx_to_lists(file_path):
    questions = []
    answers = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:  # Ensure there are at least two columns with data
                questions.append(row[0])
                answers.append(row[1])

        return questions, answers

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")

def generate_reply(generator, context, prompt, max_gen_len, temperature, top_p):

    if context:
        model_input = [RawMessage(role="system", content=context), RawMessage(role="user", content=prompt)]
    else:
        model_input = [RawMessage(role="user", content=prompt)]

    result = generator.chat_completion(model_input, max_gen_len=max_gen_len, temperature=temperature, top_p=top_p)
    
    #print("\n\n=== MODEL PROMPT ===")
    #for msg in model_input:
    #    print(f"{msg.role.capitalize()}: {msg.content}\n")

    out_message = result.generation
    #print("\n\n=== MODEL OUTPUT ===")
    #print(f"> {out_message.role.capitalize()}: {out_message.content}")
    #print("\n==================================\n")

    return out_message.content
    

def run_main(ckpt_dir: str, temperature: float = 0.6, top_p: float = 0.9, max_seq_len: int = 512, max_batch_size: int = 4, max_gen_len: Optional[int] = None, model_parallel_size: Optional[int] = None):

    generator = Llama.build(
        ckpt_dir=ckpt_dir,
        max_seq_len=max_seq_len,
        max_batch_size=max_batch_size,
        model_parallel_size=model_parallel_size,
        seed=42
    )
    
    questions, answers = read_xlsx_to_lists("llama_models/scripts/article_28_question_answer_pairs.xlsx")
    print(questions[0])

    paraphrasedQuestions = []
    for i in range(1,5):
        paraphrasedQuestion = generate_reply(generator, None, generateParaphrasePrompt(i,questions[0]), max_gen_len, temperature, top_p)
        paraphrasedQuestions.append(paraphrasedQuestion)

    print("=== PARAPHRASED QUESTIONS ===")
    for paraphrased_question in paraphrasedQuestions:
        print("\n", paraphrased_question)

    '''
    context = "i am going to test if you have a memory ok? plase remember these three words: green, lemon, house."
    prompt = "now i am going to test your memorization. waht were the three words i told you to remember?"
    generate_reply(generator, context, prompt, max_gen_len, temperature, top_p)

    
    dialogs = [
        [RawMessage(role="system", content=context),
        RawMessage(role="user", content=prompt)]
    ]

    for dialog in dialogs:
        result = generator.chat_completion(
            dialog,
            max_gen_len=max_gen_len,
            temperature=temperature,
            top_p=top_p,
        )
        print("QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQqqq")
        for msg in dialog:
            print(f"{msg.role.capitalize()}: {msg.content}\n")

        out_message = result.generation
        print(f"> {out_message.role.capitalize()}: {out_message.content}")
        print("\n==================================\n")
    '''
    

# Example usage
if __name__ == "__main__":
    print("hello test")
    
    fire.Fire(run_main)

    #if len(sys.argv) < 2:
    #    print("Usage: python script.py <xlsx_file_path>")
    #    sys.exit(1)

    #file_path = sys.argv[1]  # Get the file path from command-line arguments
    
    quit()
    

    paraphrasedQuestions = []
    for question in questions:
        generateParaphrasedQuestions(question, 5)

    

    print("Questions:", questions)
    print("Answers:", answers)
